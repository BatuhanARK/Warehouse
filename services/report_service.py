from config import get_session
from models.movement import Movement
from models.stock import Stock
from models.customer import Customer
from models.product import Product
from models.location import Location
from sqlalchemy.orm import joinedload
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm


class ReportService:
    def __init__(self):
        self.session = get_session()

    def _refresh(self):
        try:
            self.session.close()
        except:
            pass
        self.session = get_session()

    def get_movement_report(self, start_date=None, end_date=None,
                            movement_type=None, customer_id=None):
        self._refresh()
        query = self.session.query(Movement).options(
            joinedload(Movement.customer),
            joinedload(Movement.product),
            joinedload(Movement.from_loc),
            joinedload(Movement.to_loc),
        )
        if start_date:
            query = query.filter(Movement.movement_date >= start_date)
        if end_date:
            query = query.filter(Movement.movement_date <= end_date)
        if movement_type:
            query = query.filter(Movement.movement_type == movement_type)
        if customer_id:
            query = query.filter(Movement.customer_id == customer_id)
        return query.order_by(Movement.movement_date.desc()).all()

    def get_stock_report(self, customer_id=None):
        self._refresh()
        query = self.session.query(Stock).options(
            joinedload(Stock.customer),
            joinedload(Stock.product),
            joinedload(Stock.location),
        )
        if customer_id:
            query = query.filter(Stock.customer_id == customer_id)
        return query.all()

    # ── EXCEL ───────────────────────────────────────────────
    def export_movements_excel(self, movements, filepath):
        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "Hareket Raporu"

        header_fill = PatternFill("solid", fgColor="1B3A6B")
        alt_fill    = PatternFill("solid", fgColor="DBEAFE")
        headers = ["ID", "Tarih", "Tip", "Müşteri", "Ürün",
                   "Çıkış Lok.", "Giriş Lok.", "Miktar", "Referans"]

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = Font(bold=True, color="FFFFFF", size=11)
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 25

        type_colors = {"IN": "DCFCE7", "OUT": "FEE2E2", "TRANSFER": "FEF3C7"}
        for row, m in enumerate(movements, 2):
            values = [
                m.id, str(m.movement_date), m.movement_type,
                m.customer.name if m.customer else "",
                m.product.name  if m.product  else "",
                m.from_loc.code if m.from_loc  else "-",
                m.to_loc.code   if m.to_loc    else "-",
                float(m.quantity), m.reference or "",
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if row % 2 == 0:
                    cell.fill = alt_fill
                if col == 3:
                    bg = type_colors.get(str(val), "FFFFFF")
                    cell.fill = PatternFill("solid", fgColor=bg)

        widths = [8, 12, 10, 20, 25, 12, 12, 10, 15]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        wb.save(filepath)

    def export_stock_excel(self, stocks, filepath):
        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "Stok Raporu"

        header_fill = PatternFill("solid", fgColor="1B3A6B")
        alt_fill    = PatternFill("solid", fgColor="DBEAFE")
        headers = ["Müşteri", "SKU", "Ürün Adı", "Lokasyon", "Miktar", "Birim"]

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = Font(bold=True, color="FFFFFF", size=11)
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 25

        for row, s in enumerate(stocks, 2):
            values = [
                s.customer.name if s.customer else "",
                s.product.sku   if s.product  else "",
                s.product.name  if s.product  else "",
                s.location.code if s.location else "",
                float(s.quantity),
                s.product.unit  if s.product  else "",
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if row % 2 == 0:
                    cell.fill = alt_fill
                if col == 5 and isinstance(val, float) and val <= 10:
                    cell.font = Font(color="DC2626", bold=True)

        widths = [20, 15, 25, 12, 10, 8]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        wb.save(filepath)

    # ── PDF ─────────────────────────────────────────────────
    def export_movements_pdf(self, movements, filepath, title="Hareket Raporu"):
        doc      = SimpleDocTemplate(filepath, pagesize=A4,
                                     rightMargin=1.5*cm, leftMargin=1.5*cm,
                                     topMargin=2*cm, bottomMargin=2*cm)
        styles   = getSampleStyleSheet()
        elements = []

        # Başlık
        title_style = ParagraphStyle(
            "title", parent=styles["Title"],
            fontSize=16, textColor=colors.HexColor("#1B3A6B"),
            spaceAfter=20
        )
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.3*cm))

        # Tablo verisi
        headers = ["ID", "Tarih", "Tip", "Müşteri", "Ürün",
                   "Çıkış", "Giriş", "Miktar"]
        data    = [headers]

        type_colors_pdf = {
            "IN":       colors.HexColor("#16a34a"),
            "OUT":      colors.HexColor("#dc2626"),
            "TRANSFER": colors.HexColor("#d97706"),
        }

        for m in movements:
            data.append([
                str(m.id),
                str(m.movement_date),
                m.movement_type,
                m.customer.name if m.customer else "",
                m.product.name  if m.product  else "",
                m.from_loc.code if m.from_loc  else "-",
                m.to_loc.code   if m.to_loc    else "-",
                str(int(float(m.quantity))),
            ])

        col_widths = [1*cm, 2.5*cm, 2*cm, 3.5*cm, 4*cm, 2.5*cm, 2.5*cm, 2*cm]
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0),  colors.HexColor("#1B3A6B")),
            ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, 0),  9),
            ("FONTSIZE",    (0, 1), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#DBEAFE")]),
            ("GRID",        (0, 0), (-1, -1), 0.3, colors.HexColor("#e5e7eb")),
            ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING",  (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING",   (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
        ]))
        elements.append(t)
        doc.build(elements)

    def export_stock_pdf(self, stocks, filepath, title="Stok Raporu"):
        doc      = SimpleDocTemplate(filepath, pagesize=A4,
                                     rightMargin=1.5*cm, leftMargin=1.5*cm,
                                     topMargin=2*cm, bottomMargin=2*cm)
        styles   = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle(
            "title", parent=styles["Title"],
            fontSize=16, textColor=colors.HexColor("#1B3A6B"),
            spaceAfter=20
        )
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.3*cm))

        headers = ["Müşteri", "SKU", "Ürün Adı", "Lokasyon", "Miktar", "Birim"]
        data    = [headers]

        for s in stocks:
            qty = float(s.quantity)
            data.append([
                s.customer.name if s.customer else "",
                s.product.sku   if s.product  else "",
                s.product.name  if s.product  else "",
                s.location.code if s.location else "",
                str(int(qty)) if qty == int(qty) else str(qty),
                s.product.unit  if s.product  else "",
            ])

        col_widths = [3.5*cm, 2.5*cm, 5*cm, 2.5*cm, 2*cm, 1.5*cm]
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0),  colors.HexColor("#1B3A6B")),
            ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, 0),  9),
            ("FONTSIZE",    (0, 1), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#DBEAFE")]),
            ("GRID",        (0, 0), (-1, -1), 0.3, colors.HexColor("#e5e7eb")),
            ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING",  (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING",   (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
        ]))
        elements.append(t)
        doc.build(elements)